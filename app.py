from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field, ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import asyncio

app = FastAPI()

# Modelos Pydantic
class Measurement(BaseModel):
    value: float = Field(..., description="Valor da medição")
    unit: str = Field(..., description="Unidade da medição")

class SensorData(BaseModel):
    temperature: Measurement
    dissolved_oxygen: Measurement
    salinity: Measurement
    turbidity: Measurement
    microplastics: Measurement

# Caminho do arquivo Excel
EXCEL_FILE = "sensor_data.xlsx"

# Bloqueio assíncrono para acessar o arquivo Excel
file_lock = asyncio.Lock()

def initialize_excel(file_path: str):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Data"
        headers = [
            "Temperature Value", "Temperature Unit",
            "Dissolved Oxygen Value", "Dissolved Oxygen Unit",
            "Salinity Value", "Salinity Unit",
            "Turbidity Value", "Turbidity Unit",
            "Microplastics Value", "Microplastics Unit"
        ]
        ws.append(headers)
        wb.save(file_path)

@app.post("/data/")
async def receive_data(data: SensorData):
    try:
        await file_lock.acquire()
        try:
            initialize_excel(EXCEL_FILE)
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            
            row = [
                data.temperature.value, data.temperature.unit,
                data.dissolved_oxygen.value, data.dissolved_oxygen.unit,
                data.salinity.value, data.salinity.unit,
                data.turbidity.value, data.turbidity.unit,
                data.microplastics.value, data.microplastics.unit
            ]
            ws.append(row)
            wb.save(EXCEL_FILE)
        finally:
            file_lock.release()
        
        return {"message": "Data received and stored in Excel"}
    except ValidationError as ve:
        raise HTTPException(status_code=400, detail=ve.errors())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
async def read_root():
    return {"message": "API de Ingestão de Dados de Poluição de Recifes de Coral"}

@app.get("/data/")
async def list_data():
    try:
        if not os.path.exists(EXCEL_FILE):
            raise HTTPException(status_code=404, detail="No data found")
        
        await file_lock.acquire()
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            data = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                entry = {
                    "temperature": {"value": row[0], "unit": row[1]},
                    "dissolved_oxygen": {"value": row[2], "unit": row[3]},
                    "salinity": {"value": row[4], "unit": row[5]},
                    "turbidity": {"value": row[6], "unit": row[7]},
                    "microplastics": {"value": row[8], "unit": row[9]},
                }
                data.append(entry)
            
            return data
        finally:
            file_lock.release()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
